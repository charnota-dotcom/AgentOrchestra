"""Render uploaded files into a form the CLIs can consume.

* Images: pass-through (the CLI accepts a path).  We optionally resize
  oversized images to keep token cost down, but only when Pillow is
  available — falling back to raw bytes if not.
* Spreadsheets (.xlsx / .xls / .csv): rendered to one fenced markdown
  table per sheet, capped at MAX_ROWS x MAX_COLS so a 50k-row sheet
  doesn't blow the model's context.

Lazy SDK imports: openpyxl / Pillow are optional installs.  When
missing, we surface a clear AttachmentRenderError instead of crashing.
"""

from __future__ import annotations

import csv
import io
import logging
import re
from dataclasses import dataclass
from pathlib import Path

log = logging.getLogger(__name__)


# Soft caps so a single sheet doesn't dominate the prompt.  Operators
# can still see truncated rows in the file itself; this limits what
# the model receives in its context.
MAX_ROWS_PER_SHEET = 200
MAX_COLS_PER_SHEET = 30
MAX_CELL_CHARS = 200

IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".webp", ".gif"}
SPREADSHEET_EXTS = {".xlsx", ".xls", ".csv"}

# Resize cap: if the long edge is bigger than this, we shrink to fit
# with bilinear interpolation.  GIFs are skipped to preserve animation.
MAX_IMAGE_LONG_EDGE = 1600


class AttachmentRenderError(Exception):
    """Raised when we fail to render an uploaded file (bad format,
    parser missing, etc.).  The caller surfaces this as a flash on the
    upload result; the bytes are still stored on disk so the operator
    can re-try with a different agent / model later.
    """


@dataclass
class RenderResult:
    rendered_text: str | None  # None for images
    mime_type: str
    bytes_written: int


def classify_kind(path: Path) -> str:
    """Return 'image' | 'spreadsheet' | '' based on extension."""
    ext = path.suffix.lower()
    if ext in IMAGE_EXTS:
        return "image"
    if ext in SPREADSHEET_EXTS:
        return "spreadsheet"
    return ""


_SAFE_NAME = re.compile(r"[^A-Za-z0-9._-]+")


def sanitize_filename(name: str) -> str:
    """Strip path separators and weird chars so the file lands inside
    its agent dir without surprises.  Empty results fall back to
    ``upload`` so the row never has an empty stored_path.
    """
    base = Path(name).name  # strip any directory
    cleaned = _SAFE_NAME.sub("_", base).strip("._-")
    return cleaned or "upload"


def render_attachment(src: Path, *, dest: Path, kind: str) -> RenderResult:
    """Read ``src``, write a normalised copy to ``dest``, and return
    the metadata + rendered text (for spreadsheets).

    ``dest`` parent dir must already exist.
    """
    if kind == "image":
        return _render_image(src, dest)
    if kind == "spreadsheet":
        return _render_spreadsheet(src, dest)
    raise AttachmentRenderError(f"unknown attachment kind: {kind!r}")


# ---------------------------------------------------------------------------
# Images
# ---------------------------------------------------------------------------


def _mime_for_image(ext: str) -> str:
    return {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".webp": "image/webp",
        ".gif": "image/gif",
    }.get(ext.lower(), "application/octet-stream")


def _render_image(src: Path, dest: Path) -> RenderResult:
    ext = src.suffix.lower()
    mime = _mime_for_image(ext)
    raw = src.read_bytes()
    # Skip resize for animated GIFs — the first-frame extraction would
    # silently drop animation, which surprises operators.
    if ext == ".gif":
        dest.write_bytes(raw)
        return RenderResult(rendered_text=None, mime_type=mime, bytes_written=len(raw))

    try:
        from PIL import Image  # type: ignore[import-not-found]
    except ImportError:
        # No Pillow installed — store as-is.  The CLI can still accept it.
        dest.write_bytes(raw)
        return RenderResult(rendered_text=None, mime_type=mime, bytes_written=len(raw))

    # Bound the maximum decoded pixel count.  Without this, a small
    # PNG with extreme dimensions can balloon to gigabytes during
    # resize.  50 MP is generous (~7000x7000) but bounded.
    Image.MAX_IMAGE_PIXELS = 50_000_000

    try:
        with Image.open(io.BytesIO(raw)) as im:
            w, h = im.size
            # Pillow only checks MAX_IMAGE_PIXELS lazily; do an explicit
            # guard so we never attempt to resize a hostile image.
            if w * h > Image.MAX_IMAGE_PIXELS:
                log.warning(
                    "image %s pixel count %d > cap %d; storing raw bytes",
                    src.name,
                    w * h,
                    Image.MAX_IMAGE_PIXELS,
                )
                dest.write_bytes(raw)
                return RenderResult(rendered_text=None, mime_type=mime, bytes_written=len(raw))
            long_edge = max(w, h)
            if long_edge <= MAX_IMAGE_LONG_EDGE:
                dest.write_bytes(raw)
                return RenderResult(rendered_text=None, mime_type=mime, bytes_written=len(raw))
            scale = MAX_IMAGE_LONG_EDGE / long_edge
            new_size = (int(w * scale), int(h * scale))
            resized = im.resize(new_size, Image.Resampling.BILINEAR)
            buf = io.BytesIO()
            # Re-encode in the original format when possible; PNG keeps
            # transparency, JPEG wins on bytes for photos.
            fmt = im.format or "PNG"
            if fmt.upper() == "JPEG":
                if resized.mode != "RGB":
                    resized = resized.convert("RGB")
                resized.save(buf, format="JPEG", quality=85, optimize=True)
            else:
                resized.save(buf, format=fmt)
            data = buf.getvalue()
            dest.write_bytes(data)
            return RenderResult(rendered_text=None, mime_type=mime, bytes_written=len(data))
    except Exception:
        log.exception("image resize failed for %s; storing raw bytes", src.name)
        dest.write_bytes(raw)
        return RenderResult(rendered_text=None, mime_type=mime, bytes_written=len(raw))


# ---------------------------------------------------------------------------
# Spreadsheets
# ---------------------------------------------------------------------------


def _render_spreadsheet(src: Path, dest: Path) -> RenderResult:
    ext = src.suffix.lower()
    raw = src.read_bytes()
    dest.write_bytes(raw)
    try:
        if ext == ".csv":
            text = _render_csv(raw)
            mime = "text/csv"
        elif ext == ".xlsx":
            text = _render_xlsx(src)
            mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        elif ext == ".xls":
            text = _render_xls(src)
            mime = "application/vnd.ms-excel"
        else:
            raise AttachmentRenderError(f"unsupported spreadsheet ext: {ext}")
    except AttachmentRenderError:
        raise
    except Exception as exc:
        log.exception("spreadsheet render failed for %s", src.name)
        text = (
            f"_(could not render {src.name}: {exc!s}; first 4 KB of raw text follows)_\n\n"
            + _safe_decode(raw[:4096])
        )
        mime = "application/octet-stream"
    return RenderResult(rendered_text=text, mime_type=mime, bytes_written=len(raw))


def _safe_decode(b: bytes) -> str:
    return b.decode("utf-8", errors="replace")


def _render_csv(raw: bytes) -> str:
    # csv module is happy with str; decode tolerantly.
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = []
    truncated_cols = False
    for i, row in enumerate(reader):
        if i >= MAX_ROWS_PER_SHEET:
            rows.append([f"_(truncated; CSV had more rows than the {MAX_ROWS_PER_SHEET}-row cap)_"])
            break
        if len(row) > MAX_COLS_PER_SHEET:
            truncated_cols = True
            row = row[:MAX_COLS_PER_SHEET]
        rows.append([_clip_cell(c) for c in row])
    parts = ["### Sheet 1 (CSV)\n", _to_markdown_table(rows)]
    if truncated_cols:
        parts.append(f"\n_(showing first {MAX_COLS_PER_SHEET} columns)_")
    return "\n".join(parts)


def _render_xlsx(src: Path) -> str:
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except ImportError as exc:
        raise AttachmentRenderError(
            "openpyxl not installed; install agentorchestra[attachments]"
        ) from exc

    wb = load_workbook(filename=str(src), data_only=True, read_only=True)
    out: list[str] = []
    try:
        for sheet in wb.worksheets:
            rows: list[list[str]] = []
            col_capped = False
            row_iter = sheet.iter_rows(values_only=True)
            row_overflow = False
            for i, row in enumerate(row_iter):
                if i >= MAX_ROWS_PER_SHEET:
                    # Stop iterating immediately — the previous code
                    # walked the whole sheet just to compute row_total,
                    # which is a bomb on a million-row sheet.
                    row_overflow = True
                    break
                cells = list(row[:MAX_COLS_PER_SHEET])
                if len(row) > MAX_COLS_PER_SHEET:
                    col_capped = True
                rows.append([_clip_cell(c) for c in cells])
            heading = f"### Sheet: {sheet.title}"
            if not rows:
                out.append(f"{heading}\n_(empty sheet)_")
                continue
            parts = [heading, _to_markdown_table(rows)]
            if row_overflow:
                parts.append(
                    f"_(showing first {MAX_ROWS_PER_SHEET} rows; sheet has more)_",
                )
            if col_capped:
                parts.append(f"_(showing first {MAX_COLS_PER_SHEET} columns)_")
            out.append("\n".join(parts))
    finally:
        wb.close()
    return "\n\n".join(out) if out else "_(no sheets)_"


def _render_xls(src: Path) -> str:
    try:
        import xlrd  # type: ignore[import-not-found]
    except ImportError as exc:
        raise AttachmentRenderError(
            "xlrd not installed; install agentorchestra[attachments] "
            "(legacy .xls only — convert to .xlsx if possible)"
        ) from exc

    book = xlrd.open_workbook(str(src), on_demand=True)
    out: list[str] = []
    try:
        for sheet in book.sheets():
            rows: list[list[str]] = []
            for r in range(min(sheet.nrows, MAX_ROWS_PER_SHEET)):
                cols = sheet.row_values(r)[:MAX_COLS_PER_SHEET]
                rows.append([_clip_cell(c) for c in cols])
            heading = f"### Sheet: {sheet.name}"
            if not rows:
                out.append(f"{heading}\n_(empty sheet)_")
                continue
            parts = [heading, _to_markdown_table(rows)]
            if sheet.nrows > MAX_ROWS_PER_SHEET:
                parts.append(f"_(showing first {MAX_ROWS_PER_SHEET} of {sheet.nrows} rows)_")
            if sheet.ncols > MAX_COLS_PER_SHEET:
                parts.append(f"_(showing first {MAX_COLS_PER_SHEET} columns)_")
            out.append("\n".join(parts))
    finally:
        # Release the mmap so the file isn't held open on Windows.
        try:
            book.release_resources()
        except Exception:
            log.debug("xlrd release_resources failed", exc_info=True)
    return "\n\n".join(out) if out else "_(no sheets)_"


def _clip_cell(value: object) -> str:
    if value is None:
        return ""
    s = str(value).replace("|", "\\|").replace("\n", " ").strip()
    if len(s) > MAX_CELL_CHARS:
        return s[: MAX_CELL_CHARS - 1] + "…"
    return s


def _to_markdown_table(rows: list[list[str]]) -> str:
    if not rows:
        return ""
    width = max(len(r) for r in rows)
    norm = [r + [""] * (width - len(r)) for r in rows]
    header = norm[0] if norm else []
    sep = ["---"] * width
    body = norm[1:] if len(norm) > 1 else []
    lines = [
        "| " + " | ".join(header) + " |",
        "| " + " | ".join(sep) + " |",
    ]
    for r in body:
        lines.append("| " + " | ".join(r) + " |")
    return "\n".join(lines)
